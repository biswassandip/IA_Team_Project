import os
from docx import Document
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import PyPDF2
import zipfile

def search_files_with_keyword(directory_or_file, keyword, common_file_types):
    """
    Search for files containing a keyword within a directory or a specific file.

    Args:
        directory_or_file (str): The directory path or file name to search within.
        keyword (str): The keyword to search for.
        common_file_types (list): The list of common file types to search within.

    Returns:
        dict: A dictionary with 'success' and 'files' keys. 'success' indicates if the search was successful,
              and 'files' contains the list of file paths that contain the keyword.
    """
    if os.path.isdir(directory_or_file):
        files = get_all_files(directory_or_file)
    elif os.path.isfile(directory_or_file):
        files = [directory_or_file]
    else:
        return {'success': False, 'files': []}

    found_files = search_keyword_in_files(files, keyword, common_file_types)
    return {'success': True, 'files': found_files}

def get_all_files(directory):
    """
    Get a list of all files within a directory and its subdirectories.

    Args:
        directory (str): The directory path.

    Returns:
        list: A list of file paths.
    """
    file_paths = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            file_paths.append(file_path)
    return file_paths

def search_keyword_in_files(files, keyword, common_file_types):
    """
    Search for a keyword within a list of files.

    Args:
        files (list): A list of file paths.
        keyword (str): The keyword to search for.
        common_file_types (list): The list of common file types to search within.

    Returns:
        list: A list of file paths that contain the keyword.
    """
    found_files = []
    for file_path in files:
        file_ext = os.path.splitext(file_path)[1].lower()

        # Skip image files
        if file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
            continue

        if file_ext in common_file_types:
            if file_ext == '.txt' or file_ext == '.html' or file_ext == '.css':
                if search_keyword_in_text_file(file_path, keyword):
                    found_files.append(file_path)
            elif file_ext == '.docx':
                if search_keyword_in_docx_file(file_path, keyword):
                    found_files.append(file_path)
            elif file_ext in ['.xls', '.xlsx']:
                if search_keyword_in_excel_file(file_path, keyword):
                    found_files.append(file_path)
            elif file_ext == '.xml':
                if search_keyword_in_xml_file(file_path, keyword):
                    found_files.append(file_path)
            elif file_ext == '.pdf':
                if search_keyword_in_pdf_file(file_path, keyword):
                    found_files.append(file_path)
            elif file_ext == '.zip':
                if search_keyword_in_zip_file(file_path, keyword, common_file_types):
                    found_files.append(file_path)

    return found_files

def search_keyword_in_text_file(file_path, keyword):
    """
    Search for a keyword within a text file.

    Args:
        file_path (str): The path of the text file.
        keyword (str): The keyword to search for.

    Returns:
        bool: True if the keyword is found, False otherwise.
    """
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        if keyword in f.read():
            return True
    return False

def search_keyword_in_docx_file(file_path, keyword):
    """
    Search for a keyword within a DOCX file.

    Args:
        file_path (str): The path of the DOCX file.
        keyword (str): The keyword to search for.

    Returns:
        bool: True if the keyword is found, False otherwise.
    """
    document = Document(file_path)
    for paragraph in document.paragraphs:
        if keyword in paragraph.text:
            return True
    return False

def search_keyword_in_excel_file(file_path, keyword):
    """
    Search for a keyword within an Excel file.

    Args:
        file_path (str): The path of the Excel file.
        keyword (str): The keyword to search for.

    Returns:
        bool: True if the keyword is found, False otherwise.
    """
    workbook = load_workbook(file_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and keyword in str(cell.value):
                    return True
    return False

def search_keyword_in_xml_file(file_path, keyword):
    """
    Search for a keyword within an XML file.

    Args:
        file_path (str): The path of the XML file.
        keyword (str): The keyword to search for.

    Returns:
        bool: True if the keyword is found, False otherwise.
    """
    tree = ET.parse(file_path)
    root = tree.getroot()
    if keyword in ET.tostring(root, encoding='utf-8').decode('utf-8'):
        return True
    return False

def search_keyword_in_pdf_file(file_path, keyword):
    """
    Search for a keyword within a PDF file.

    Args:
        file_path (str): The path of the PDF file.
        keyword (str): The keyword to search for.

    Returns:
        bool: True if the keyword is found, False otherwise.
    """
    with open(file_path, 'rb') as f:
        pdf_reader = PyPDF2.PdfFileReader(f)
        for page_num in range(pdf_reader.numPages):
            page = pdf_reader.getPage(page_num)
            page_text = page.extract_text()
            if keyword in page_text:
                return True
    return False

def search_keyword_in_zip_file(file_path, keyword, common_file_types):
    """
    Search for a keyword within a ZIP file.

    Args:
        file_path (str): The path of the ZIP file.
        keyword (str): The keyword to search for.
        common_file_types (list): The list of common file types to search within.

    Returns:
        bool: True if the keyword is found, False otherwise.
    """
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        for file_info in zip_ref.infolist():
            if not file_info.is_dir() and not file_info.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp')):
                with zip_ref.open(file_info) as f:
                    content = f.read()
                    if keyword in content.decode('utf-8', errors='ignore'):
                        return True
    return False

# Example usage
root_directory = '/Users/gini/00-downloads-sorted/docs'  # Replace with your root directory or file name
search_keyword = 'shop'  # Replace with your desired keyword
common_file_types = ['.txt', '.doc', '.docx', '.xls', '.xlsx', '.xml', '.pdf', '.png', '.zip']

abc = search_files_with_keyword(root_directory,search_keyword,common_file_types)
# print(f"{abc.found_files}")