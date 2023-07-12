import os
from docx import Document
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from PyPDF2 import PdfReader
import zipfile
import sys
import common.general_handlers.constants as cons


class SearchInFiles:


    def __init__(self, config, dir_or_file, keyword):
        """
        Initialize the SearchInFiles class with configuration values from the config ini file.
            dir_or_file (str): The directory path or file name to search within.
            keyword (str): The keyword to search for.

        Args:
            config (object): The config file
        """
        self.config = config
        self.ini_file_path = self.config.get(cons.INI_HEADER_GENERAL, cons.KEY_INI_FILE_PATH)


        self.directory_or_file = dir_or_file
        self.keyword = keyword
        self.common_file_types = [ext.strip() for ext in self.config[cons.INI_HEADER_SEARCH_IN_FILE_TYPES][cons.KEY_INCLUDE].split(',')]
        self.skip_file_types = [ext.strip() for ext in self.config[cons.INI_HEADER_SEARCH_IN_FILE_TYPES][cons.KEY_INCLUDE].split(',')] 


    def search_files_with_keyword(self):
        
        """
        Search for files containing a keyword within a directory or a specific file.

        Returns:
            dict: A dictionary with 'files' and 'errors' keys. 'files' gives the list of file paths that has the keyword,
                and 'errors' contains the list error messages with file paths.
        """
        if os.path.isdir(self.directory_or_file):
            files = self.get_all_files(self.directory_or_file)
        elif os.path.isfile(self.directory_or_file):
            files = [self.directory_or_file]
        else:
            return {'files': [], 'errors':[]}

        return self.search_keyword_in_files(files, self.keyword, self.common_file_types, self.skip_file_types)


    def get_all_files(self,directory):
        """
        Get a list of all files within a directory and its subdirectories.

        Args:
            directory (str): The directory path.

        Returns:
            list: A list of file paths.
        """
        file_paths = []
        for root, dirs, files in os.walk(directory):
            for file in files:
                file_path = os.path.join(root, file)
                file_paths.append(file_path)
        return file_paths


    def search_keyword_in_files(self, files, keyword, common_file_types, skip_file_types):
        """
        Search for a keyword within a list of files.

        Args:
            files (list): A list of file paths.
            keyword (str): The keyword to search for.
            common_file_types (list): The list of common file types to search within.
            skip_file_types (list): The list of file types to skip.

        Returns:
            list: A list of file paths that contain the keyword.
        """
        found_files = []
        errors = []
        for file_path in files:
            file_ext = os.path.splitext(file_path)[1].lower()

            # Skip specified file types
            if skip_file_types and file_ext in skip_file_types:
                continue
            
            if file_ext in common_file_types:

                if file_ext == '.txt' or file_ext == '.html' or file_ext == '.css':
                    results = self.search_keyword_in_text_file(file_path, keyword)
                elif file_ext == '.docx':
                    results = self.search_keyword_in_docx_file(file_path, keyword)
                elif file_ext in ['.xls', '.xlsx']:
                    results = self.search_keyword_in_excel_file(file_path, keyword)
                elif file_ext == '.xml':
                    results = self.search_keyword_in_xml_file(file_path, keyword)
                elif file_ext == '.pdf':
                    results = self.search_keyword_in_pdf_file(file_path, keyword)
                elif file_ext == '.zip':
                    results = self.search_keyword_in_zip_file(file_path, keyword, common_file_types, skip_file_types)

                success = results["success"]

                if success:
                    found_files.append(file_path)
                elif (not success) :
                    error_message= results['error_message'].strip() + ""
                    if (len(error_message)>0):
                        errors.append(error_message)

        return {'files': found_files, 'errors': errors}


    def search_keyword_in_text_file(self, file_path, keyword):
        """
        Search for a keyword within a text file.

        Args:
            file_path (str): The path of the text file.
            keyword (str): The keyword to search for.

        Returns:
            bool: True if the keyword is found, False otherwise.
        """
        b_return = False
        error_message = ""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                if keyword in f.read():
                    b_return = True
        except:
            # Handle decryption error
            error_message = str(sys.exc_info()[1]) + " " + file_path

            b_return = False

        return {"success":b_return, "file":file_path, "error_message":error_message}


    def search_keyword_in_docx_file(self, file_path, keyword):
        """
        Search for a keyword within a DOCX file.

        Args:
            file_path (str): The path of the DOCX file.
            keyword (str): The keyword to search for.

        Returns:
            bool: True if the keyword is found, False otherwise.
        """
        b_return = False
        error_message = ""
        try:
            document = Document(file_path)
            for paragraph in document.paragraphs:
                if keyword in paragraph.text:
                    b_return = True
                    break
        except:
            # Handle decryption error
            error_message = str(sys.exc_info()[1]) + " " + file_path

            b_return = False

        return {"success":b_return, "file":file_path, "error_message":error_message}


    def search_keyword_in_excel_file(self, file_path, keyword):
        """
        Search for a keyword within an Excel file.

        Args:
            file_path (str): The path of the Excel file.
            keyword (str): The keyword to search for.

        Returns:
            bool: True if the keyword is found, False otherwise.
        """
        b_return = False
        error_message = ""
        try:
            workbook = load_workbook(file_path)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and keyword in str(cell.value):
                            b_return =  True
                            break
                        if b_return: break
                    if b_return: break
                if b_return: break
        except:
            # Handle decryption error
            error_message = str(sys.exc_info()[1]) + " " + file_path

            b_return = False

        return {"success":b_return, "file":file_path, "error_message":error_message}


    def search_keyword_in_xml_file(self, file_path, keyword):
        """
        Search for a keyword within an XML file.

        Args:
            file_path (str): The path of the XML file.
            keyword (str): The keyword to search for.

        Returns:
            bool: True if the keyword is found, False otherwise.
        """
        b_return = False
        error_message = ""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            if keyword in ET.tostring(root, encoding='utf-8').decode('utf-8'):
                b_return = True
        except:
            # Handle decryption error
            error_message = str(sys.exc_info()[1]) + " " + file_path

            b_return = False

        return {"success":b_return, "file":file_path, "error_message":error_message}


    def search_keyword_in_pdf_file(self, file_path, keyword):
        """
        Search for a keyword within a PDF file.

        Args:
            file_path (str): The path of the PDF file.
            keyword (str): The keyword to search for.

        Returns:
            bool: True if the keyword is found, False otherwise.
        """
        b_return = False
        error_message = ""
        try:
            pdf = PdfReader(file_path)
            for page in pdf.pages:
                if keyword in page.extract_text():
                    b_return = True
                    break
        except:
            # Handle decryption error
            error_message = str(sys.exc_info()[1]) + " " + file_path

            b_return = False

        return {"success":b_return, "file":file_path, "error_message":error_message}


    def search_keyword_in_zip_file(self, file_path, keyword, common_file_types, skip_file_types):
        """
        Search for a keyword within a ZIP file.

        Args:
            file_path (str): The path of the ZIP file.
            keyword (str): The keyword to search for.
            common_file_types (list): The list of common file types to search within.
            skip_file_types (list): The list of file types to skip.

        Returns:
            bool: True if the keyword is found, False otherwise.
        """
        b_return = False
        error_message = ""
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                for file_info in zip_ref.infolist():
                    if not file_info.is_dir() and not file_info.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp')):
                        with zip_ref.open(file_info) as f:
                            content = f.read()
                            if keyword in content.decode('utf-8', errors='ignore'):
                                return True
        except:
            # Handle decryption error
            error_message = str(sys.exc_info()[1]) + " " + file_path

            b_return = False

        return {"success":b_return, "file":file_path, "error_message":error_message}