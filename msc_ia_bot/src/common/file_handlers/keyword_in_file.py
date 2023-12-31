import os
from docx import Document
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from PyPDF2 import PdfReader
import zipfile
import sys
from common.general_handlers.utils import Utils

"""
**class: KeywordInFile**

This class is responsible for find specific keyword within defined file types.
The supported file types as of now are:
".txt, .doc, .docx, .xls, .xlsx, .xml, .pdf, .png, .zip, .html, .css"
"""

class KeywordInFile:


    def __init__(self):
        pass

    def search_keyword_in_files(self, file_path, keyword, common_file_types):

        """
        Search for a keyword within a list of files.

        Args:
            files (list): A list of file paths.
            keyword (str): The keyword to search for.
            common_file_types (list): The list of common file types to search within.
            skip_file_types (list): The list of file types to skip.

        Returns:
            dict: A dictionary with 'files' and 'errors' keys. 'files' gives the list of file paths that has the keyword,
                and 'errors' contains the list error messages with file paths.
        """

        file_ext = os.path.splitext(file_path)[1].lower()
        results = False

        if file_ext in common_file_types:

            if file_ext == '.txt' or file_ext == '.html' or file_ext == '.css':
                results = self.search_keyword_in_text_file(file_path, keyword)
            elif file_ext == '.docx':
                results = self.search_keyword_in_docx_file(file_path, keyword)
            elif file_ext in ['.xls', '.xlsx']:
                results = self.search_keyword_in_excel_file(file_path, keyword)
            elif file_ext == '.xml':
                results = self.search_keyword_in_xml_file(file_path, keyword)
            elif file_ext == '.pdf':
                results = self.search_keyword_in_pdf_file(file_path, keyword)
            elif file_ext == '.zip':
                results = self.search_keyword_in_zip_file(file_path, keyword)

        return results


    def search_keyword_in_text_file(self, file_path, keyword):

        """
        Search for a keyword within a text file.

        Args:
            file_path (str): The path of the text file.
            keyword (str): The keyword to search for.

        Returns:
            dict: A dictionary with 'success', 'file' and 'error_message' keys.
        """

        b_return = False

        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                if keyword in f.read():
                    b_return = True
            
        except:
            Utils.error_message(str(sys.exc_info()[1]) + " " + file_path)
            b_return = False

        return b_return


    def search_keyword_in_docx_file(self, file_path, keyword):
        
        """
        Search for a keyword within a DOCX file.

        Args:
            file_path (str): The path of the DOCX file.
            keyword (str): The keyword to search for.

        Returns:
            dict: A dictionary with 'success', 'file' and 'error_message' keys.
        """

        b_return = False
        
        try:
            document = Document(file_path)
            for paragraph in document.paragraphs:
                if keyword in paragraph.text:
                    b_return = True
                    break
        except:
            Utils.error_message(str(sys.exc_info()[1]) + " " + file_path)
            b_return = False

        return b_return


    def search_keyword_in_excel_file(self, file_path, keyword):

        """
        Search for a keyword within an Excel file.

        Args:
            file_path (str): The path of the Excel file.
            keyword (str): The keyword to search for.

        Returns:
            dict: A dictionary with 'success', 'file' and 'error_message' keys.
        """

        b_return = False

        try:
            workbook = load_workbook(file_path)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and keyword in str(cell.value):
                            b_return =  True
                            break
                        if b_return: break
                    if b_return: break
                if b_return: break
        except:
            Utils.error_message(str(sys.exc_info()[1]) + " " + file_path)
            b_return = False

        return b_return


    def search_keyword_in_xml_file(self, file_path, keyword):

        """
        Search for a keyword within an XML file.

        Args:
            file_path (str): The path of the XML file.
            keyword (str): The keyword to search for.

        Returns:
            dict: A dictionary with 'success', 'file' and 'error_message' keys.
        """

        b_return = False

        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            if keyword in ET.tostring(root, encoding='utf-8').decode('utf-8'):
                b_return = True
        except:
            Utils.error_message(str(sys.exc_info()[1]) + " " + file_path)
            b_return = False

        return b_return


    def search_keyword_in_pdf_file(self, file_path, keyword):

        """
        Search for a keyword within a PDF file.

        Args:
            file_path (str): The path of the PDF file.
            keyword (str): The keyword to search for.

        Returns:
            dict: A dictionary with 'success', 'file' and 'error_message' keys.
        """

        b_return = False

        try:
            pdf = PdfReader(file_path)
            for page in pdf.pages:
                if keyword in page.extract_text():
                    b_return = True
                    break
        except:
            Utils.error_message(str(sys.exc_info()[1]) + " " + file_path)
            b_return = False

        return b_return


    def search_keyword_in_zip_file(self, file_path, keyword):

        """
        Search for a keyword within a ZIP file.

        Args:
            file_path (str): The path of the ZIP file.
            keyword (str): The keyword to search for.

        Returns:
            dict: A dictionary with 'success', 'file' and 'error_message' keys.
        """

        b_return = False

        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                for file_info in zip_ref.infolist():
                    if not file_info.is_dir() and not file_info.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp')):
                        with zip_ref.open(file_info) as f:
                            content = f.read()
                            if keyword in content.decode('utf-8', errors='ignore'):
                                return True
        except:
            Utils.error_message(str(sys.exc_info()[1]) + " " + file_path)
            b_return = False

        return b_return
    

if __name__ == '__main__':
    print(f"")